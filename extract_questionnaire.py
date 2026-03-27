import json
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl  # type: ignore


@dataclass
class QuestionSummary:
    code: str
    text: str
    answer_option_count: int


@dataclass
class ScopeSummary:
    code: str
    label: str
    questions: List[QuestionSummary]


@dataclass
class ThemeSummary:
    title: str
    scope_items: List[ScopeSummary]


def slugify(value: str) -> str:
    return (
        "".join(ch.lower() if ch.isalnum() else "-" for ch in value)
        .strip("-")
        .replace("--", "-")
    )


def load_sheet(path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    return wb[sheet_name]


def row_to_dict(row) -> Dict[str, Any]:
    cols = list("ABCDEFGHIJKLMNOP")
    data: Dict[str, Any] = {}
    for cell in row:
        coord = getattr(cell, "coordinate", None)
        if not coord:
            continue
        col_letter = "".join(ch for ch in coord if ch.isalpha())
        if col_letter in cols:
            data[col_letter] = cell.value
    # ensure all keys exist
    for c in cols:
        data.setdefault(c, None)
    return data


def main() -> None:
    root = Path(__file__).parent
    wb_path = root / "GreenDIGIT-D8.1-Self-Assessment-Questionnaire.xlsx"
    sheet_name = "2-3. Assessment & Actions"
    ws = load_sheet(wb_path, sheet_name)

    # Find header row (where B == "Topic area")
    header_row_idx: Optional[int] = None
    for row in ws.iter_rows(min_row=1, max_row=40):
        data = row_to_dict(row)
        if str(data.get("B", "")).strip() == "Topic area":
            header_row_idx = row[0].row  # actual row index
            break

    if header_row_idx is None:
        raise SystemExit("Could not find header row with 'Topic area' in column B.")

    start_row = header_row_idx + 1
    max_row = ws.max_row

    themes_order: List[str] = []
    theme_scopes: Dict[str, Dict[str, ScopeSummary]] = defaultdict(dict)

    # Structures for normalized static configuration
    themes_config: Dict[str, Dict[str, Any]] = {}  # key: theme_title
    scope_config: Dict[str, Dict[str, Any]] = {}   # key: scope_id
    questions_config: Dict[str, Dict[str, Any]] = {}  # key: question_id

    issues = {
        "ambiguousMappings": [],
        "missingValues": [],
        "repeatedCodes": [],
        "inconsistencies": [],
    }  # type: Dict[str, List[str]]

    current_theme_title: Optional[str] = None
    current_scope_label: Optional[str] = None
    current_scope_code: Optional[str] = None

    question_codes_seen: Dict[str, Tuple[str, str]] = {}

    total_questions = 0
    questions_with_exactly_three_options = 0

    r = start_row
    while r <= max_row:
        row = ws[r]
        data = row_to_dict(row)

        b = (data.get("B") or "").strip() if isinstance(data.get("B"), str) else data.get("B")
        d = data.get("D")
        e = data.get("E")
        m = data.get("M")

        # Theme row: B has value, D/E/M empty (and B is not header label)
        if b and not d and not e and not m and str(b).strip() != "Topic area":
            current_theme_title = str(b).strip()
            if current_theme_title not in themes_order:
                themes_order.append(current_theme_title)
                # create theme config entry (id is slug of title)
                themes_config[current_theme_title] = {
                    "id": slugify(current_theme_title),
                    "title": current_theme_title,
                }
            # reset scope when theme changes
            current_scope_label = None
            current_scope_code = None
            r += 1
            continue

        # Question row: D is requirement code, M == 1 denotes first row of 3-row block
        if d and m == 1:
            if current_theme_title is None:
                issues["ambiguousMappings"].append(
                    f"Question code {d} at row {r} has no current theme."
                )

            # Scope label may appear only on first question of the scope
            if isinstance(b, str) and b.strip():
                current_scope_label = b.strip()
                # derive scope code from prefix before ':', if present
                if ":" in current_scope_label:
                    current_scope_code = current_scope_label.split(":", 1)[0].strip()
                else:
                    current_scope_code = current_scope_label
            # If still no scope from previous context, record ambiguity
            if current_scope_label is None or current_scope_code is None:
                issues["ambiguousMappings"].append(
                    f"Question code {d} at row {r} has no scope label in column B."
                )

            question_code = str(d).strip()
            question_text = str(e).strip() if isinstance(e, str) else str(e or "")

            theme_key = current_theme_title or "UNKNOWN_THEME"
            scope_code = current_scope_code or "UNKNOWN_SCOPE"
            scope_label = current_scope_label or "UNKNOWN_SCOPE"

            theme_id = themes_config.get(theme_key, {}).get(
                "id", slugify(theme_key)
            )
            scope_id = scope_code  # use scope code (e.g. "G1") as stable ID

            # Track repeated question codes
            if question_code in question_codes_seen:
                prev_theme, prev_scope = question_codes_seen[question_code]
                issues["repeatedCodes"].append(
                    f"Question code {question_code} repeated at row {r}; "
                    f"previously under theme '{prev_theme}', scope '{prev_scope}'."
                )
            else:
                question_codes_seen[question_code] = (theme_key, scope_code)

            # Ensure scope summary exists
            scopes_for_theme = theme_scopes[theme_key]
            if scope_code not in scopes_for_theme:
                scopes_for_theme[scope_code] = ScopeSummary(
                    code=scope_code,
                    label=scope_label,
                    questions=[],
                )
            # Ensure scope config exists
            if scope_id not in scope_config:
                scope_config[scope_id] = {
                    "id": scope_id,
                    "themeId": theme_id,
                    "code": scope_code,
                    "label": scope_label,
                }

            # Collect 3-row block for answer options
            answer_rows: List[Dict[str, Any]] = []
            for offset in range(3):
                row_idx = r + offset
                if row_idx > max_row:
                    issues["inconsistencies"].append(
                        f"Question code {question_code} at row {r} truncated; "
                        f"expected 3 rows but sheet ended early."
                    )
                    break
                row_data = row_to_dict(ws[row_idx])
                answer_rows.append(row_data)

            total_questions += 1

            # Validate answer rows count and M values
            option_scores = [ar.get("M") for ar in answer_rows]
            if len(answer_rows) == 3 and option_scores == [1, 2, 3]:
                questions_with_exactly_three_options += 1
            else:
                issues["inconsistencies"].append(
                    f"Question code {question_code} at row {r} has unexpected "
                    f"capability levels in column M: {option_scores}."
                )

            # Build static Question config entry
            q_row = answer_rows[0] if answer_rows else {}
            lifecycle_phase = q_row.get("F")
            env_dimension = q_row.get("G")
            assoc_metric = q_row.get("H")
            reference = q_row.get("I")

            # Build answerOptions array from three rows
            answer_options: List[Dict[str, Any]] = []
            for ar in answer_rows:
                answer_options.append(
                    {
                        "score": int(ar.get("M")),
                        "description": ar.get("N"),
                        "impactMagnitude": ar.get("J"),
                        "impactLikelihood": ar.get("K"),
                        "recommendedAction": ar.get("O"),
                    }
                )

            questions_config[question_code] = {
                "id": question_code,
                "themeId": theme_id,
                "scopeId": scope_id,
                "code": question_code,
                "text": question_text,
                "lifecyclePhase": lifecycle_phase,
                "environmentalDimension": env_dimension,
                "associatedMetric": assoc_metric,
                "reference": reference,
                "answerOptions": answer_options,
            }

            # Check required question-level fields (D,E,F,G,H,I) on first row
            for col, name in [
                ("D", "code"),
                ("E", "text"),
                ("F", "lifecyclePhase"),
                ("G", "environmentalDimension"),
                ("H", "associatedMetric"),
                ("I", "reference"),
            ]:
                # Only D and E are strictly required; others are optional per spec,
                # but we still note if they are entirely missing.
                if col in ("D", "E") and not q_row.get(col):
                    issues["missingValues"].append(
                        f"Question code {question_code} row {r} missing required {name} in column {col}."
                    )

            # Check required answer-option-level fields from J,K,M,N,O on each row
            for idx, ar in enumerate(answer_rows):
                level = ar.get("M")
                for col, name in [
                    ("J", "impactMagnitude"),
                    ("K", "impactLikelihood"),
                    ("M", "score"),
                    ("N", "description"),
                    ("O", "recommendedAction"),
                ]:
                    if ar.get(col) in (None, ""):
                        issues["missingValues"].append(
                            f"Question code {question_code} level {level} "
                            f"(row {r+idx}) missing {name} in column {col}."
                        )

            # Append question summary (compact text)
            short_text = question_text
            if len(short_text) > 80:
                short_text = short_text[:77] + "..."

            scopes_for_theme[scope_code].questions.append(
                QuestionSummary(
                    code=question_code,
                    text=short_text,
                    answer_option_count=len(answer_rows),
                )
            )

            # Move to next block after these 3 rows
            r += 3
            continue

        # Any other row: just advance
        r += 1

    # Build ordered theme summaries
    theme_summaries: List[ThemeSummary] = []
    for theme_title in themes_order:
        scopes_dict = theme_scopes.get(theme_title, {})
        # preserve scope ordering by sorted scope code for determinism
        scope_items = [
            scopes_dict[code] for code in sorted(scopes_dict.keys(), key=str)
        ]
        theme_summaries.append(
            ThemeSummary(
                title=theme_title,
                scope_items=scope_items,
            )
        )

    # Build static questionnaire configuration object
    themes_array = [
        themes_config[title] for title in themes_order if title in themes_config
    ]

    scope_items_array = list(scope_config.values())

    questions_array = list(questions_config.values())

    questionnaire_config = {
        "themes": themes_array,
        "scopeItems": scope_items_array,
        "questions": questions_array,
        "validation": {
            "totalQuestions": total_questions,
            "questionsWithExactlyThreeOptions": questions_with_exactly_three_options,
            "allQuestionsHaveThreeOptions": total_questions
            == questions_with_exactly_three_options,
            "issues": issues,
        },
    }

    # Ensure output directory exists: src/lib/saq
    out_dir = root / "src" / "lib" / "saq"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "questionnaire.data.json"

    with out_path.open("w", encoding="utf-8") as f:
        json.dump(
            {
                # Persist only the immutable config fields the app needs.
                "themes": questionnaire_config["themes"],
                "scopeItems": questionnaire_config["scopeItems"],
                "questions": questionnaire_config["questions"],
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    # Also print a compact summary to stdout for inspection.
    summary = {
        "themeCount": len(themes_array),
        "scopeItemCount": len(scope_items_array),
        "questionCount": len(questions_array),
        "validation": questionnaire_config["validation"],
    }

    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()

