import json
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:
    print(json.dumps({"error": "missing_openpyxl"}))
    raise


def main() -> None:
    root = Path(__file__).parent
    wb_path = root / "GreenDIGIT-D8.1-Self-Assessment-Questionnaire.xlsx"

    wb = openpyxl.load_workbook(wb_path, data_only=True)
    sheets = wb.sheetnames

    result: dict = {"sheets": sheets, "sample": {}}

    target_name = "2-3. Assessment & Actions"
    if target_name not in wb:
        print(json.dumps(result, indent=2))
        return

    ws = wb[target_name]

    # Capture first 40 non-empty rows, columns A–P.
    max_rows = 40
    cols = list("ABCDEFGHIJKLMNOP")

    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows):
        row_dict = {}
        for cell in row:
            # Some merged cells come through as MergedCell without column_letter;
            # use the cell coordinate to derive the column instead.
            coord = getattr(cell, "coordinate", None)
            if not coord:
                continue
            # coordinate is like "A1", "B2", etc. Strip row digits to get the column.
            col_letter = "".join(ch for ch in coord if ch.isalpha())
            if col_letter in cols:
                row_dict[col_letter] = cell.value
        if any(v not in (None, "") for v in row_dict.values()):
            rows_data.append(row_dict)

    result["sample"][target_name] = rows_data

    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()

